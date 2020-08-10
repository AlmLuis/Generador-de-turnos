####importacion de modulos####

from openpyxl import Workbook, load_workbook, cell
import random, getpass, sys, smtplib, socket

########################################################################
####acceder a archivos####

data = load_workbook( 'turnos_proyecto.xlsx' )
turnos_disponibles = data[ 'TURNOS DISPONIBLES' ]
turnos_otorgados = data[ 'TURNOS OTORGADOS' ]

data1 = load_workbook( filename = 'usuarios_registrados.xlsx' )
hoja_usuarios = data1[ 'USUARIOS' ]


########################################################################
####funciones####

def listado():
	lista = []
	for x in range( 2, 12 ):
		profesionales = turnos_disponibles.cell( row = x, column = 1 ).value
		especialidad = turnos_disponibles.cell( row = x, column = 2 ).value
		lista.append( [profesionales, especialidad] )
	return lista

def horarios():
	lista = []
	for x in range( 3, 19 ):
		turnos = turnos_disponibles.cell( row = 2, column = x ).value
		lista.append( turnos )
	return lista

def turnos( especialista, horarios ):
	lista = []
	for e in especialista:
		lista.append( [e, horarios] )
	return lista
	
def registro_usuarios():
	data1 = load_workbook( filename = 'usuarios_registrados.xlsx' )
	hoja_usuarios = data1[ 'USUARIOS' ]
	lista_usuarios = []
	usuario = raw_input( 'Ingrese nombre y apellido: ' )
	email=raw_input( "Ingrese cuenta de mail(si no posee ingrese 'n'): " )
	contrasenia = raw_input( 'Ingrese su nueva contrasenia: ' )
	confirmacion = raw_input( 'Ingrese nuevamente su contrasenia: ' )
	if contrasenia == confirmacion:
		if email == 'n':
			email = 'No posee'
		hoja_usuarios.append( [ usuario, contrasenia, email ] )
		lista_usuarios.append( [ usuario, contrasenia, email ] )
		data1.save( filename = 'usuarios_registrados.xlsx' )
		data.save( filename = 'turnos_proyecto.xlsx' )
		print ('Su registro se ha efectuado en forma satisfactoria')
		raw_input( '' )
	else:
		print ('Datos incorrectos, ingreselos nuevamente')
		usuario = raw_input( 'Ingrese nombre y apellido: ' )
		email=raw_input( "Ingrese cuenta de mail(si no posee ingrese 'n'): " )
		contrasenia = raw_input( 'Ingrese su nueva contrasenia: ' )
		confirmacion = raw_input( 'Ingrese nuevamente su contrasenia: ' )
		if contrasenia == confirmacion:
			if email == 'n':
					email = 'No posee'
			hoja_usuarios.append( [usuario, contrasenia, email] )
			lista_usuarios.append( [ usuario, contrasenia, email ] )
			data1.save( filename = 'usuarios_registrados.xlsx' )
			data.save( filename = 'turnos_proyecto.xlsx' )
			print ('Muchas gracias por registrarse')
			raw_input('')
		else:
			print ('Datos incorrectos, ultimo intento')
			usuario = raw_input( 'Ingrese nombre y apellido: ' )
			email=raw_input( "Ingrese cuenta de mail(si no posee ingrese 'n'): " )
			contrasenia = raw_input( 'Ingrese su nueva contrasenia: ' )
			confirmacion = raw_input( 'Ingrese nuevamente su contrasenia: ' )
			if contrasenia == confirmacion:
				if email == 'n':
					email = 'No posee'
				hoja_usuarios.append( [usuario, contrasenia, email] )
				lista_usuarios.append( [ usuario, contrasenia, email ] )
				data1.save( 'usuarios_registrados.xlsx' )
				data.save( 'turnos_proyecto.xlsx' )
				print ('Muchas gracias por registrarse')
				raw_input( '' )
			else:
				print ('Ha alcanzado el limite de ingreso de datos para el registro de usuario')
				raw_input( '' )
	data1.save( 'usuarios_registrados.xlsx' )
	data.save( 'turnos_proyecto.xlsx' )
	return lista_usuarios
	
def lista_usuarios():
		usuarios = []
		for filas in hoja_usuarios.iter_cols( min_col=1, max_col=1, min_row=2 ):
			for fila in filas:
				valor = fila.value
				usuarios.append( valor )
			return usuarios

def lista_contrasenias():
		contrasenias = []
		for filas in hoja_usuarios.iter_cols( min_col=2, max_col=2, min_row=2 ):
			for fila in filas:
				valor = fila.value
				contrasenias.append( valor )
			return contrasenias
			
def lista_email():
	email = []
	for filas in hoja_usuarios.iter_cols( min_col=3, max_col=3, min_row=2 ):
		for fila in filas:
			valor = fila.value
			email.append( valor )
		return email

def lista_de_usuarios_registrados( usuarios, contrasenias, email ):
	lista = [ usuarios, contrasenias, email ]
	return lista
	
def comparo( lista, usuario, contrasenia ):
	usuarios = lista[0]
	contrasenias = lista[1]
	if usuario in usuarios:
		posicion = usuarios.index( usuario )
		contrasenia_correspondiente = contrasenias[ posicion ]
		if contrasenia == contrasenia_correspondiente:
			return 'Contrasenia correcta'
		else:
			contrasenia = getpass.getpass( 'Contrasenia incorrecta\nIngrese contrasenia: ' )
			if contrasenia == contrasenia_correspondiente:
				return 'Contrasenia correcta'
			else:
				contrasenia = getpass.getpass( 'Contrasenia incorrecta\nUltimo intento\nIngrese contrasenia: ' )
				if contrasenia == contrasenia_correspondiente:
					return 'Contrasenia correcta'
				else:
					return 'Contrasenia incorrecta'
	
def lista_horarios_de_turnos( lista_turnos_disponibles ):
	for turno in lista_turnos_disponibles:
		print (turno)

def especialistas_listado( lista_especialistas ):
	especialistas = []
	for e in lista_especialistas:
		prof = e [ 1 ]
		especialistas.append( prof )
	return especialistas

def especialidades_listado( lista_especialistas ):
	especialidades = []
	for e in lista_especialistas:
		esp = e [ 0 ]
		especialidades.append( esp )
	return especialidades
	
def registro_de_turno( usuario,lista_turnos_disponibles,esp,prof ):
	listas_de_turnos = []
	horario_elegido = input( 'Ingrese el numero de opcion del horario deseado: ' )
	hora = lista_turnos_disponibles[ horario_elegido - 1 ]
	especialidad_solicitada = input( 'Ingrese el numero de opcion correspondiente a especialidad por la cual quiera atenderse: ' )
	servicio = esp[ especialidad_solicitada - 1 ]
	if hora in lista_turnos_disponibles and servicio in esp:
		numero_de_turno = random.randint( 1000, 9999 )
		fecha_de_turno = raw_input( 'Ingrese la fecha para la cual desea el turno: ' )
		posicion = esp.index( servicio )
		especialista_solicitado = prof[ posicion ]
		confirma = raw_input( 'Confirma que desea solicitar el turno:(s/n) ' )
		if confirma == 's':
			turnos_otorgados.append( [numero_de_turno, fecha_de_turno, hora, servicio, especialista_solicitado, usuario] )
			data1.save( 'usuarios_registrados.xlsx' )
			data1.close()
			data.save( 'turnos_proyecto.xlsx' )
			data.close()
			listas_de_turnos.append( [numero_de_turno, fecha_de_turno, hora, servicio, especialista_solicitado, usuario] )
			print ('Su numero de tramite es ' + str( numero_de_turno ))
			print ('Su turno se ha reservado de forma exitosa')
			print
		elif confirma == 'n':
			print ('Solicitud cancelada exitosamente')
			print
		else:
			print ('Opcion incorrecta')
			print
	else:
		print ('Datos erroneos o inexistentes')
		c = raw_input( 'Desea continuar(s/n): ' )
		if c == 's' :
			horario_elegido = input( 'Ingrese el numero de opcion del horario deseado: ' )
			hora = lista_turnos_disponibles[ horario_elegido - 1 ]
			especialidad_solicitada = input( 'Ingrese el numero de opcion correspondiente a especialidad por la cual quiera atenderse: ' )
			servicio = esp[ especialidad_solicitada - 1 ]
			if hora in lista_turnos_disponibles and servicio in esp:
				numero_de_turno = random.randint( 1000, 2000 )
				fecha_de_turno = raw_input( 'Ingrese la fecha para la cual desea el turno: ' )
				posicion = esp.index( servicio )
				especialista_solicitado = prof[posicion]
				confirma = raw_input( 'Confirma que desea solicitar el turno:(s/n) ' )
				if confirma == 's':
					turnos_otorgados.append( [numero_de_turno, fecha_de_turno, hora, servicio, especialista_solicitado, usuario] )
					data1.save( 'usuarios_registrados.xlsx' )
					data1.close()
					data.save( 'turnos_proyecto.xlsx' )
					data.close()
					listas_de_turnos.append( [numero_de_turno, fecha_de_turno, hora, servicio, especialista_solicitado, usuario] )
					print ('Su numero de tramite es ' + str( numero_de_turno ))
					print ('Su turno se ha reservado de forma exitosa')
					print
				elif confirma == 'n':
					print ('Solicitud cancelada exitosamente')
					print
				else:
					print ('Opcion incorrecta')
					print
		elif c == 'n':
			print ('Muchas gracias por su visita')
		else:
			'Opcion incorrecta'
	return listas_de_turnos
		
def cancela_turno( usuario,turnos ):
	for rows in turnos.iter_rows( ):
		row = list( rows )
		for r in row:
			u = r.value
			posicion = row.index( r )
			if u == usuario:
				r.value = 'Turno cancelado'
	data1.save( 'usuarios_registrados.xlsx' )
	data1.close( )
	data.save( 'turnos_proyecto.xlsx' )
	data.close( )
	print ('Turno cancelado exitosamente')
	print

def consulta_turnos( tramite, turnos ):
	lista = []
	for filas in turnos.iter_rows( min_row = 2 ):
		numero = filas[0].value
		fecha = filas[1].value
		hora = filas[2].value
		esp = filas[3].value
		prof = filas[4].value
		paciente = filas[5].value
		if numero == tramite:
			lista.append( [numero, fecha, hora, esp, prof, paciente] )
	return lista
	
def tabla_consulta_turnos( listas ):
	for lista in listas:
		numero = lista[0]
		fecha = lista[1]
		hora = lista[2]
		esp = lista[3]
		prof = lista[4]
		paciente = lista[5] 	
		print ('Para el numero de tramite ' + str( numero ) + ' se registran los siguientes datos:')
		print ('* Fecha: ' + str( fecha ))
		print ('* Horario: ' + str( hora ))
		print ('* Perteneciente a: ' + paciente)
		print ('* Servicio de: ' + esp)
		print ('* Profesional de la salud: ' + prof)
		print
	return [numero, fecha, hora, esp, prof, paciente]
	
def tabla( lista ):
	print ('Los horarios disponibles son los siguientes: ')
	orden = 0
	for elemento in lista:
		orden = orden + 1
		print (str( orden ) + ' - ' + str( elemento ))
	print

def tabla_elem( lista ):
	orden = 0
	for x in lista:
		orden = orden + 1
		print (str( orden ) + ' - ' + str( x ))

def envia_correo( lista, usuario, mensaje ):
	usuarios = lista [0]
	emails = lista [2]
	if usuario in usuarios:
		 posicion = usuarios.index ( usuario )
		 email = emails [ posicion ]
		 # Conexion con el servidor 
		 try:
		  smtpserver = smtplib.SMTP( "smtp.gmail.com", 587 )
		  smtpserver.ehlo()
		  smtpserver.starttls()
		  smtpserver.ehlo()
		  # Datos 
		  try:
		   gmail_user = 'clinicasdelsur.unaj2018@gmail.com'
		   gmail_pwd = 'unaj2018'
		   smtpserver.login( gmail_user, gmail_pwd )
		  except smtplib.SMTPException:
		   smtpserver.close()		 
		 except ( socket.gaierror, socket.error, socket.herror, smtplib.SMTPException ), e:
		  print ('Fallo en la conexion con Gmail')
		  print (getpass.getpass( 'Presione ENTER para continuar '))	 
		 while True:
		  to = email
		  if to != "" :
		   break
		  else :
		   print ('El correo es necesario!!!')		 
		 sub = 'Servicio de turnos online, Clinicas del Sur'
		 bodymsg = mensaje
		 header = "Para: " + to +"\n" + "De: " + gmail_user + "\n" + "Asunto: " + sub + "\n"
		 msg = header + "\n" + bodymsg + "\n\n"		 
		 try:
		  smtpserver.sendmail( gmail_user, to, msg )
		 except smtplib.SMTPException :
		  print ("El correo no pudo ser enviado" + "\n")
		  smtpserver.close()
		 smtpserver.close()
		
########################################################################
####programa####

lista_profesionales = listado()		
horarios_disponibles = horarios()
turnos_por_especialidad = turnos( lista_profesionales, horarios_disponibles )

lista_de_usuarios = lista_usuarios()
lista_de_contrasenias = lista_contrasenias()
lista_de_emails = lista_email()
lista_general_de_usuarios_y_contrasenias = lista_de_usuarios_registrados( lista_de_usuarios, lista_de_contrasenias, lista_de_emails )

profesionales = especialistas_listado( lista_profesionales )
especialidades = especialidades_listado( lista_profesionales )

########################################################################
####menu inicial####
while True:
	print ('=' * 240)
	print ( 'Bienvenido al portal informatico de Clinicas del sur\n\n\t\tMENU PRINCIPAL\n\nA continuacion seleccione la opcion deseada:' )
	print 
	print ('1 - Ingreso\n2 - Registro\n3 - Salir')
	print
	menu = input( 'Opcion: ' )
	if menu == 1 :
		while True :
			usuario = raw_input( 'Ingrese usuario (nombre y apellido) : ' )
			contrasenia = getpass.getpass( 'Ingrese su contrasenia: ' )
			comprobacion = comparo( lista_general_de_usuarios_y_contrasenias, usuario, contrasenia )
			if comprobacion == 'Contrasenia correcta' :
				while True :
				#submenu de seccion de accion referente al turno
					print ('*' * 120)
					print
					print ('Bienvenido ' + str( usuario ) + '.')
					print
					print ('A continuacion seleccione la opcion deseada: ')
					print
					print ('1 - Solicitud de turno\n2 - Cancelacion de turno\n3 - Consulte su turno\n4 - Volver')
					print
					opcion_submenu = input( 'Opcion: ' )
					print
					if opcion_submenu == 1 :
						while True :
							####menu de solicitud####
							lista_de_horarios = horarios()
							tabla( lista_de_horarios )
							lista_profesionales_especialidades = listado()
							prof = especialistas_listado( lista_profesionales_especialidades )
							esp = especialidades_listado( lista_profesionales_especialidades )
							print
							print ('Servicios disponibles: ')
							print
							tabla_elem( esp )
							print
							r = registro_de_turno( usuario, lista_de_horarios, esp, prof )
							if r != [] :
								tramite_numero = r [ 0 ][ 0 ]
								consulta_realizada = consulta_turnos( tramite_numero, turnos_otorgados )
								t = tabla_consulta_turnos( consulta_realizada )
								fecha = t [ 1 ]
								hora = t [ 2 ]
								paciente = t [ 5 ]
								esp = t [ 3 ]
								prof = t [ 4 ]
								mensaje = 'Este es un mail de confirmacion de solicitud de turno\nSu solicitud se encuentra bajo el numero de tramite ' + str( tramite_numero ) + ' con los siguientes datos:\n* Fecha: ' + str( fecha ) + '\n* Horario: ' + str( hora ) + '\n* Perteneciente a: ' + paciente + '\n* Servicio de: ' + esp + '\n* Profesional de la salud: ' + prof + '\nMuchas gracias por elegirnos'
								envia_correo(lista_general_de_usuarios_y_contrasenias, usuario, mensaje)
								volver = raw_input( 'Volver al menu anterior(s/n): ' )
								if volver == 's' :
									break
								elif volver == 'n':
									print ('Gracias por su visita')
									sys.exit()
								elif volver != 's' and volver != 'n':
									print ('Opcion incorrecta')
									sys.exit()
							else:
								break
					elif opcion_submenu == 2 :
						while True :
							####menu de cancelacion####
							cancela_turno( usuario, turnos_otorgados )
							mensaje = 'Este es el mail de confirmacion de cancelacion de turno,\na traves del mismo Clinicas del Sur le informa\nque la cancelacion se ha efectuado de forma exitosa.\nMuchas gracias por elegirnos.'
							envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
							volver = raw_input( 'Volver al menu anterior(s/n): ' )
							if volver == 's' :
								break
							elif volver == 'n' :
								print ('Gracias por su visita')
								sys.exit()
							elif volver != 's' and volver != 'n' :
								print ('Opcion incorrecta')
								sys.exit()
					elif opcion_submenu == 3 :
						while True :
							####menu de visualizacion de turnos####
							tramite = input( 'Ingrese su numero de tramite: ' )
							consulta_realizada = consulta_turnos( tramite, turnos_otorgados )
							tabla_consulta_turnos( consulta_realizada )
							volver = raw_input( 'Volver al menu anterior(s/n): ' )
							if volver == 's' :
								break
							elif volver == 'n' :
								print ('Gracias por su visita')
								sys.exit()
							elif volver != 's' and volver != 'n' :
								print ('Opcion incorrecta')
								sys.exit()
					elif opcion_submenu == 4 :
						break
					else :
						continua = raw_input( 'Opcion incorrecta\nDesea continuar?(s/n): ' )
						if continua == 's' :
							seleccion = input( 'Seleccione la opcion deseada: ' )
							if seleccion == 1 :
								while True :
									####menu de solicitud####
									lista_de_horarios = horarios()
									tabla( lista_de_horarios )
									lista_profesionales_especialidades = listado()
									prof = especialistas_listado( lista_profesionales_especialidades )
									esp = especialidades_listado( lista_profesionales_especialidades )
									print
									print ('Servicios disponibles: ')
									print
									tabla_elem( esp )
									print
									r = registro_de_turno( usuario, lista_de_horarios, esp, prof )
									tramite_numero = r [ 0 ][ 0 ]
									consulta_realizada = consulta_turnos( tramite_numero, turnos_otorgados )
									t = tabla_consulta_turnos( consulta_realizada )
									fecha = t [ 1 ]
									hora = t [ 2 ]
									paciente = t [ 5 ]
									esp = t [ 3 ]
									prof = t [ 4 ]
									mensaje = 'Este es un mail de confirmacion de solicitud de turno\nSu solicitud se encuentra bajo el numero de tramite ' + str( tramite_numero ) + ' con los siguientes datos:\n* Fecha: ' + str( fecha ) + '\n* Horario: ' + str( hora ) + '\n* Perteneciente a: ' + paciente + '\n* Servicio de: ' + esp + '\n* Profesional de la salud: ' + prof + '\nMuchas gracias por elegirnos'
									envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
									volver = raw_input( 'Volver al menu anterior(s/n): ' )
									if volver == 's' :
										break
									elif volver == 'n' :
										print ('Gracias por su visita')
										sys.exit()
									elif volver != 's' and volver != 'n' :
										print ('Opcion incorrecta')
										sys.exit()
							elif seleccion == 2 :
								while True :
									####menu de cancelacion####
									cancela_turno( usuario, turnos_otorgados )
									mensaje = 'Este es el mail de confirmacion de cancelacion de turno,\na traves del mismo Clinicas del Sur le informa\nque la cancelacion se ha efectuado de forma exitosa.\nMuchas gracias por elegirnos.'
									envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
									volver = raw_input( 'Volver al menu anterior(s/n): ' )
									if volver == 's' :
										break
									elif volver == 'n' :
										print ('Gracias por su visita')
										sys.exit()
									elif volver != 's' and volver != 'n' :
										print ('Opcion incorrecta')
										sys.exit()
							elif seleccion == 3 :
								while True :
									####menu de visualizacion de turnos####
									tramite = input( 'Ingrese su numero de tramite: ' )
									consulta_realizada = consulta_turnos( tramite, turnos_otorgados )
									tabla_consulta_turnos( consulta_realizada )
									volver = raw_input( 'Volver al menu anterior(s/n): ' )
									if volver == 's' :
										break
									elif volver == 'n' :
										print ('Gracias por su visita')
										sys.exit()
									elif volver != 's' and volver != 'n' :
										print ('Opcion incorrecta')
										sys.exit()
							else :
								print ('Opcion incorrecta')
						elif continua == 'n' :
							print ('Muchas gracias por su visita')
							sys.exit()
						else :
							'Opcion incorrecta'
							volver = raw_input( 'Volver al menu anterior(s/n): ' )
							if volver == 's' :
								break
							elif volver == 'n' :
								print ('Gracias por su visita')
								sys.exit()
							elif volver != 's' and volver != 'n' :
								print ('Opcion incorrecta')
								sys.exit()
			else :
				print ('Nombre de usuario o contrasenia incorrectos')
				print
				volver = raw_input( 'Volver al menu anterior(s/n): ' )
				if volver == 's' :
					break
				elif volver == 'n' :
					print ('Gracias por su visita')
					sys.exit()
				elif volver != 's' and volver != 'n' :
					print ('Opcion incorrecta')
					sys.exit()
			break
	elif menu == 2 :
		while True:
			lista = registro_usuarios()
			mensaje = 'Este es el mail de confirmacion de registro de usuario,\na traves del mismo Clinicas del Sur le informa\nque el registro se ha efectuqado de forma exitosa.\nA partir de este momento puede disfrutar de nuestros servicos online.\nMuchas gracias por elegirnos.'
			usuario = lista [0]
			envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
			volver = raw_input( 'Desea registrar un nuevo usuario(s/n): ' )
			if volver == 's' :
				break
			else :
				print ('Gracias por su visita')
				sys.exit()
	elif menu == 3 :
		print ('Muchas gracias por su visita')
		sys.exit()
	else :
		print ('Opcion incorrecta')
		opcion = raw_input( 'Desea seleccionar una opcion?(s/n): ' )
		if opcion == 's' :
			opcion1 = input( 'Ingrese la opcion deseada: ' )
			if opcion1 == 1 :
				usuario = raw_input( 'Ingrese usuario (nombre y apellido) : ' )
				contrasenia = getpass.getpass( 'Ingrese su contrasenia: ' )
				comprobacion = comparo( lista_general_de_usuarios_y_contrasenias, usuario, contrasenia )
				if comprobacion == 'Contrasenia correcta' :
					while True :
				#submenu de seccion de accion referente al turno
						print ('*' * 120)
						print
						print ('Bienvenido ' + str( usuario ) + '.')
						print
						print ('A continuacion seleccione la opcion deseada: ')
						print
						print ('1 - Solicitud de turno\n2 - Cancelacion de turno\n3 - Consulte su turno\n4 - Volver')
						print
						opcion_submenu = input( 'Opcion: ' )
						print
						if opcion_submenu == 1 :
							while True :
								####menu de solicitud####
								lista_de_horarios = horarios()
								tabla( lista_de_horarios )
								lista_profesionales_especialidades = listado()
								prof = especialistas_listado( lista_profesionales_especialidades )
								esp = especialidades_listado( lista_profesionales_especialidades )
								print
								print ('Servicios disponibles: ')
								print
								tabla_elem( esp )
								print
								r = registro_de_turno( usuario, lista_de_horarios, esp, prof )
								tramite_numero = r [ 0 ][ 0 ]
								consulta_realizada = consulta_turnos( tramite_numero, turnos_otorgados )
								t = tabla_consulta_turnos( consulta_realizada )
								fecha = t [ 1 ]
								hora = t [ 2 ]
								paciente = t [ 5 ]
								esp = t [ 3 ]
								prof = t [ 4 ]
								mensaje = 'Este es un mail de confirmacion de solicitud de turno\nSu solicitud se encuentra bajo el numero de tramite ' + str( tramite_numero ) + ' con los siguientes datos:\n* Fecha: ' + str( fecha ) + '\n* Horario: ' + str( hora ) + '\n* Perteneciente a: ' + paciente + '\n* Servicio de: ' + esp + '\n* Profesional de la salud: ' + prof + '\nMuchas gracias por elegirnos'
								envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
								volver = raw_input( 'Volver al menu anterior(s/n): ' )
								if volver == 's' :
									break
								elif volver == 'n' :
									print ('Gracias por su visita')
									sys.exit()
								elif volver != 's' and volver != 'n' :
									print ('Opcion incorrecta')
									sys.exit()
						elif opcion_submenu == 2 :
							while True :
								####menu de cancelacion####
								cancela_turno( usuario, turnos_otorgados )
								mensaje = 'Este es el mail de confirmacion de cancelacion de turno,\na traves del mismo Clinicas del Sur le informa\nque la cancelacion se ha efectuado de forma exitosa.\nMuchas gracias por elegirnos.'
								envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
								volver = raw_input( 'Volver al menu anterior(s/n): ' )
								if volver == 's' :
									break
								elif volver == 'n' :
									print ('Gracias por su visita')
									sys.exit()
								elif volver != 's' and volver != 'n' :
									print ('Opcion incorrecta')
									sys.exit()
						elif opcion_submenu == 3 :
							while True :
								####menu de visualizacion de turnos####
								tramite = input( 'Ingrese su numero de tramite: ' )
								consulta_realizada = consulta_turnos( tramite, turnos_otorgados )
								tabla_consulta_turnos( consulta_realizada )
								volver = raw_input( 'Volver al menu anterior(s/n): ' )
								if volver == 's' :
									break
								elif volver == 'n' :
									print ('Gracias por su visita')
									sys.exit()
								elif volver != 's' and volver != 'n' :
									print ('Opcion incorrecta')
									sys.exit()
						elif opcion_submenu == 4 :
							break
						else :
							continua = raw_input( 'Opcion incorrecta\nDesea continuar?(s/n): ' )
							if continua == 's' :
								seleccion = input( 'Seleccione la opcion deseada: ' )
								if seleccion == 1 :
									while True :
										####menu de solicitud####
										lista_de_horarios = horarios()
										tabla( lista_de_horarios )
										lista_profesionales_especialidades = listado()
										prof = especialistas_listado( lista_profesionales_especialidades )
										esp = especialidades_listado( lista_profesionales_especialidades )
										print
										print ('Servicios disponibles: ')
										print
										tabla_elem( esp )
										print
										r = registro_de_turno( usuario, lista_de_horarios, esp, prof )
										tramite_numero = r [ 0 ][ 0 ]
										consulta_realizada = consulta_turnos( tramite_numero, turnos_otorgados )
										t = tabla_consulta_turnos( consulta_realizada )
										fecha = t [ 1 ]
										hora = t [ 2 ]
										paciente = t [ 5 ]
										esp = t [ 3 ]
										prof = t [ 4 ]
										mensaje = 'Este es un mail de confirmacion de solicitud de turno\nSu solicitud se encuentra bajo el numero de tramite ' + str( tramite_numero ) + ' con los siguientes datos:\n* Fecha: ' + str( fecha ) + '\n* Horario: ' + str( hora ) + '\n* Perteneciente a: ' + paciente + '\n* Servicio de: ' + esp + '\n* Profesional de la salud: ' + prof + '\nMuchas gracias por elegirnos'
										envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
										volver = raw_input( 'Volver al menu anterior(s/n): ' )
										if volver == 's' :
											break
										elif volver == 'n' :
											print ('Gracias por su visita')
											sys.exit()
										elif volver != 's' and volver != 'n' :
											print ('Opcion incorrecta')
											sys.exit()
								elif seleccion == 2 :
									while True :
										####menu de cancelacion####
										cancela_turno( usuario, turnos_otorgados )
										mensaje = 'Este es el mail de confirmacion de cancelacion de turno,\na traves del mismo Clinicas del Sur le informa\nque la cancelacion se ha efectuado de forma exitosa.\nMuchas gracias por elegirnos.'
										envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
										volver = raw_input( 'Volver al menu anterior(s/n): ' )
										if volver == 's' :
											break
										elif volver == 'n' :
											print ('Gracias por su visita')
											sys.exit()
										elif volver != 's' and volver != 'n' :
											print ('Opcion incorrecta')
											sys.exit()
								elif seleccion == 3 :
									while True :
										####menu de visualizacion de turnos####
										tramite = input( 'Ingrese su numero de tramite: ' )
										consulta_realizada = consulta_turnos( tramite, turnos_otorgados )
										tabla_consulta_turnos( consulta_realizada )
										volver = raw_input('Volver al menu anterior(s/n): ')
										if volver == 's' :
											break
										elif volver == 'n':
											print ('Gracias por su visita')
											sys.exit()
										elif volver != 's' and volver != 'n':
											print ('Opcion incorrecta')
											sys.exit()
								else :
									print ('Opcion incorrecta')
							elif continua == 'n' :
								print ('Muchas gracias por su visita')
								sys.exit()
							else :
								'Opcion incorrecta'
								volver = raw_input( 'Volver al menu anterior(s/n): ' )
								if volver == 's' :
									break
								elif volver == 'n' :
									print ('Gracias por su visita')
									sys.exit()
								elif volver != 's' and volver != 'n' :
									print ('Opcion incorrecta')
									sys.exit()
				else :
					print ('Nombre de usuario o contrasenia incorrectos\nMuchas gracias por su visita')
					raw_input( '' )
			elif opcion1 == 2 :
				registro_usuarios()
				mensaje = 'Este es el mail de confirmacion de registro de usuario,\na traves del mismo Clinicas del Sur le informa\nque el registro se ha efectuqado de forma exitosa.\nA partir de este momento puede disfrutar de nuestros servicos online.\nMuchas gracias por elegirnos.'
				envia_correo( lista_general_de_usuarios_y_contrasenias, usuario, mensaje )
				volver = raw_input( 'Desea registrar un nuevo usuario(s/n): ' )
				if volver == 's' :
					restart_program()
				else :
					print ('Gracias por su visita')
					sys.exit()
			else :
				print ('Opcion incorrecta')
				raw_input( '' )
		elif opcion == 'n' :
			print ('Muchas gracias por su visita')
			raw_input( '' )
		

########################################################################

